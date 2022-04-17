# Licensed to AC. Electrical, LLC
# Located in  1861 Marion Court Beloit, Wisconsin
# This file is owned and copyrighted by  AC. Electrical, LLC
# The SFC licenses this file
#
#  Author: Jorge Jesus Jurado-Garica
#  Title: Product Search specialist
#   Project Description: Python Script for price listing
#           and price increase for Menards for electrical products.
#           It will grab price value for products and store then into an array
#  Date of Creation: 11/11/2021
#  Rev:
#

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from website_dict import website

SKU = []
description = []
unit_price = []

# Creates the driver and it is using chrome driver for this
# we can use firefox,windows, and others if needed or wanted
driver = webdriver.Chrome(executable_path="C:\drivers\chromedriver_win32\chromedriver.exe")
# Goes to main page
def CRB_function():
    driver.get(website.get('CRB'))
    driver.maximize_window()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[3]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value) + 1)
    for x in follow_loop:
        y = x
        if(x > 37):
            y = x -36
        if (x == 37):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.CSS_SELECTOR, ".btn.btn-circle")
            driver.execute_script("arguments[0].click();", next_arrow)
            y = x -36

        xpath = "/html/body/div[5]/section/div[5]/div/div[2]/div[3]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def MBB_function():
    driver.get(website.get('MBB'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[1]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value) + 1)
    for x in follow_loop:
        y = x
        if( x > 37):
            y = x-36
        if (x == 37):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.CSS_SELECTOR, ".btn.btn-circle")
            driver.execute_script("arguments[0].click();", next_arrow)
            y = x-36
        xpath = "/html/body/div[5]/section/div[3]/div/div[2]/div[2]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def MLB_function():
    driver.get(website.get('MLB'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[1]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value) + 1)
    for x in follow_loop:
        y = x
        if( x > 37):
            y = x-36
        if (x == 37):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.XPATH, "./html/body/div[5]/section/div[3]/div/div[2]/div[2]/div[2]/div[3]/a")
            driver.execute_script("arguments[0].click();", next_arrow)
            y = x-36
        xpath = "/html/body/div[5]/section/div[3]/div/div[2]/div[2]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def SSD_function():
    driver.get(website.get('SSD'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[1]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value) + 1)
    for x in follow_loop:
        y = x
        if( x > 37):
            y = x-36
        if( x == 35):
            y = x+1
        if (x == 37):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.CSS_SELECTOR, ".btn.btn-circle")
            driver.execute_script("arguments[0].click();", next_arrow)
            y = x-36
        xpath = "/html/body/div[5]/section/div[3]/div/div[2]/div[3]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def FUS_function():
    driver.get(website.get('FUS'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[3]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value) + 1)
    for x in follow_loop:
        y = x
        if( x > 37):
            y = x-36
        if (x == 37):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.CSS_SELECTOR, ".btn.btn-circle")
            driver.execute_script("arguments[0].click();", next_arrow)
            y = x-36
        xpath = "/html/body/div[5]/section/div[5]/div/div[2]/div[3]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def FS_function():
    driver.get(website.get('F&S'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[3]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value)+1)
    for x in follow_loop:
        y = x
        if( x > 216):
            y = x-216
        elif( x > 180):
            y = x-180
        elif( x > 144):
            y = x-144
        elif( x > 108):
            y = x-108
        elif( x > 72):
            y = x-72
        elif( x > 36):
            y = x-36
        else:
            print("")
        if ( ((x % 37) == 0) ):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[5]/div/div[2]/div[3]/div[2]/div[3]/a")
            driver.execute_script("arguments[0].click();", next_arrow)

        xpath = "/html/body/div[5]/section/div[5]/div/div[2]/div[3]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
def EBX_function():
    driver.get(website.get('EBX'))
    driver.maximize_window()
    driver.delete_all_cookies()
    # gets the amount of items in the search bar
    max_value = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[1]/div[1]/div[1]/div/span[2]")
    max_value = max_value.text
    print("Number of products:", max_value)
    # for loop to read the product name and descriptions

    # product = driver.find_element(By.CSS_SELECTOR, ".search-item")
    # product = product.text
    # print(product)
    des = "/div/div/a/div[2]/div[2]/span"
    pr = "/div/div/div[4]/div/span/span/span"
    sp = "/div/div/div[3]/span"
    des_path = "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark"
    follow_loop = range(1, int(max_value)+1)
    for x in follow_loop:
        y = x
        if( x > 216):
            y = x-216
        elif( x > 180):
            y = x-180
        elif( x > 144):
            y = x-144
        elif( x > 108):
            y = x-108
        elif( x > 72):
            y = x-72
        elif( x > 36):
            y = x-36
        else:
            print("")
        if ( ((x % 37) == 0) ):
            driver.delete_all_cookies()
            next_arrow = driver.find_element(By.XPATH, "/html/body/div[5]/section/div[3]/div/div[2]/div[3]/div[2]/div[3]/a")
            driver.execute_script("arguments[0].click();", next_arrow)
        xpath = "/html/body/div[5]/section/div[3]/div/div[2]/div[3]/div[1]/div["
        xpath += str(y)
        xpath += "]"
        xpath += des
        description.append(driver.find_element(By.XPATH, xpath).text)
        xpath2 = xpath.replace(des, '')
        xpath2 += pr
        unit_price.append(driver.find_element(By.XPATH, xpath2).text)
        xpath3 = xpath2.replace(pr, '')
        xpath3 += sp
        SKU.append(driver.find_element(By.XPATH, xpath3).text)
CRB_function()
driver.delete_all_cookies()
MBB_function()
driver.delete_all_cookies()
MLB_function()
driver.delete_all_cookies()
SSD_function()
driver.delete_all_cookies()
FUS_function()
driver.delete_all_cookies()
FS_function()
driver.delete_all_cookies()
EBX_function()
# loads workbook of excel file named data
# and looks into worksheet label DV_Sentry_460-480VAC
wb = load_workbook(filename='data.xlsx')
sh = wb["Sheet1"]

for j in range(len(description)):
    sh.cell(row=j + 2, column=2).value = description[j]
    sh.cell(row=j + 2, column=5).value = unit_price[j]
    sh.cell(row=j + 2, column =1).value = SKU[j]
wb.save(filename='data.xlsx')

#description = [i.text for i in driver.find_elements(By.CSS_SELECTOR, "#search-items > .search-item .details > .pricing-info.pt-3" )]
#for description
#description = [i.text for i in driver.find_elements(By.CSS_SELECTOR, "#search-items > .search-item .details > a > .row.pt-5.pb-sm-5 > .multilines-3.text-truncate-multilines.xs-single-col-8.col-12 > .font-weight-bold.text-dark")]

